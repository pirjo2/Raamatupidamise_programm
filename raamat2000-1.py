import PySimpleGUI as sg
#import xlwt
#import xlsxwriter
#import openpyxl
from openpyxl import Workbook
#from xlwt import Workbook
from openpyxl import load_workbook
wb = Workbook()
sg.theme('LightBlue6')  # akna värvilahenduse muutmine

def arvuta(bruto, sots, pens, tooandja, tootaja, tulumaks, toolisemaksud, neto, maksud, tooandjamaks, tulum, tmv, aasta, pens2, tootaja2, sots2, tooandja2):
    bruto = float(bruto)
    aasta = bruto * 12
    if sots == True:    # sotsiaalmaksu arvutamine
        if bruto * 0.33 > 178.2:
            tooandjamaks += bruto * 0.33
            maksud += bruto * 0.33
            sots2 = bruto * 0.33
        else:    # sotsiaalmaksu minimaalne kohustus on 178.2 eurot kuus
            tooandjamaks += 178.2
            maksud += 178.2
            sots2 = 178.2
    if pens == True:    # pensioni arvutamine
        neto += bruto * 0.02
        maksud += bruto * 0.02
        toolisemaksud += bruto * 0.02
        pens2 = bruto * 0.02
    if tooandja == True:    # töötuskindlustusmakse arvutamine, tööandja
        tooandjamaks += bruto * 0.008
        maksud += bruto * 0.008
        tooandja2 = bruto * 0.008
    if tootaja == True:    # töötuskindlustusmakse arvutamine, töötaja
        neto += bruto * 0.016
        maksud += bruto * 0.016
        toolisemaksud += bruto * 0.016
        tootaja2 = bruto * 0.016
    if tulumaks == True:    # tulumaksuvaba summa arvutamine
        if aasta <= 6000:
            tmv = 500
        if aasta > 6000 and aasta <= 14400:
            tmv = 500
        if aasta > 14400 and aasta <= 25200:
            tmv = 6000 - 6000 / 10800 * (aasta - 14400)
        if aasta > 25200:
            tmv = 0
    if aasta >= 6000:    # tulumaksu arvutamine
        tulum = 0
    if aasta > 6000 and aasta <= 14400:
        tulum = (aasta - tmv - toolisemaksud * 12 - 6000) * 0.2
    if aasta > 14400 and aasta <= 25200:
        tulum = (aasta - tmv - toolisemaksud * 12) * 0.2
    if aasta > 25200:
        tulum = (aasta - toolisemaksud) * 0.2
    tulum = round(tulum / 12, 2)    # ümardamine ja lisaarvutused
    neto = round(bruto - neto - tulum, 2)
    maksud += tulum
    tooandjamaks += bruto
    maksud = round(maksud, 2)
    tooandjamaks = round(tooandjamaks, 2)
    tulum = round(tulum, 2)
    tmv = round(tmv, 2)
    toolisemaksud = round(toolisemaksud, 2)
    return [neto, maksud, tooandjamaks, tulum, tmv, toolisemaksud, pens2, tootaja2, sots2, tooandja2]

toolisemaksud = 0.0    # maksud, mis lähevad töölise bruto palgast maha
neto = 0.0    # bruto - maksud = neto, selle saab puhtalt kätte
maksud = 0.0    # kõik maksud kokku
tooandjamaks = 0.0    # tööandja kohustus maksta kõik kokku
tulum = 0.0    # tulumaks
tmv = 0.0    # tulumaksuvaba
aasta = 0.0    # aastatulu
pens2 = 0.0    # pension
tootaja2 = 0.0    # töötaja töötuskindlustusmakse
sots2 = 0.0    #sotsiaalmakse
tooandja2 = 0.0    # tööandja töötuskindlustusmakse


neto3 = 0.0    #Excelisse kirjutamiseks
pension3 = 0.0    #Excelisse kirjutamiseks
tootuskindlustustootaja3 = 0.0    #Excelisse kirjutamiseks
tulumaks3 = 0.0    #Excelisse kirjutamiseks
tootuskindlustustooandja3 = 0.0    #Excelisse kirjutamiseks
sots3 = 0.0    #Excelisse kirjutamiseks
kokku3 = 0.0    #Excelisse kirjutamiseks
lehenimi =""
aa = ""
# file = open
layout = [
    [sg.Text('PALGAKALKULAATOR'), sg.Text(size=(16,1), key = 'tekstisilt1')],
    [sg.Text('Sisestage töötaja nimi: '), sg.Text(size=(12,1), key = 'nimesilt1'),
    sg.InputText('Eesnimi', size = (12,1), do_not_clear = True, key = 'nimelahter1'),
    sg.InputText('Perenimi', size = (12,1), do_not_clear = True, key = 'nimelahter2')],
    [sg.Text('Sisestage tööandja nimi: '), sg.Text(size=(11,1), key = 'nimesilt2'),
    sg.InputText('Eesnimi', size = (12,1), do_not_clear = True, key = 'nimelahter3'),
    sg.InputText('Perenimi', size = (12,1), do_not_clear = True, key = 'nimelahter4')],
    [sg.Text('Sisestage kuu ja aasta: '), sg.Text(size=(12,1), key = 'kuupaevasilt1'),
    sg.InputText('Kuu', size = (12,1), do_not_clear = True, key = 'kuupaevalahter1'),
    sg.InputText('Aasta', size = (12,1), do_not_clear = True, key = 'kuupaevalahter2')],
    [sg.Text('Vali maksud: '), sg.Text(size=(12,1), key = 'tekstisilt2')],
    [sg.Checkbox('Maksuvaba tulu', default = True, key = 'tulumaks')],
    [sg.Checkbox('Sotsiaalmaks', default = True, key = 'sotsmaks'), sg.Checkbox('Kogumispension', default = True, key = 'pens')],
    [sg.Text('Töötuskindlustusmaksed:'), sg.Text(size=(12,1), key = 'tekstisilt3')],
    [sg.Checkbox('Tööandja', default = True, key = 'tooandja'), sg.Checkbox('Töötaja', default = True, key = 'tootaja')],
    [sg.Text('Sisestage bruto palk: '), sg.Text(size=(16,1), key = 'tekstisilt4'),
    sg.InputText('EUR', size = (9,1), do_not_clear = True, key = 'bruto')],
    [sg.Button('Kalkuleeri', key = 'button')],
    [sg.Text('TÖÖTAJA:'), sg.Text(size=(12,1))],
    [sg.Text('Palk netos: '), sg.Text(size=(12,1), key = 'neto2')],
    [sg.Text('Kogumispension: '), sg.Text(size=(12,1), key = 'pension')],
    [sg.Text('Töötuskindlustusmaks (töötaja): '), sg.Text(size=(12,1), key = 'tootuskindlustus2')],
    [sg.Text('Tulumaks: '), sg.Text(size=(12,1), key = 'tulumaks2')],
    [sg.Text('TÖÖANDJA:'), sg.Text(size=(12,1))],
    [sg.Text('Tööandja kulud kokku: '), sg.Text(size=(12,1), key = 'kokku2')],
    [sg.Text('Sotsiaalmaks: '), sg.Text(size=(12,1), key = 'sots2')],
    [sg.Text('Töötuskindlustusmaks (tööandja): '), sg.Text(size=(12,1), key = 'tooandja3')],
    [sg.Button('Lisa andmed faili', key = 'button1'),
    sg.Exit('Välju')]
    ]

window = sg.Window('Raamatupidamine 2000', layout)

while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == 'Välju':
        break
    if event == 'button':    # kui klikatakse "Kalkuleeri nuppu
        list = arvuta(values['bruto'], values['sotsmaks'], values['pens'], values['tooandja'], values['tootaja'], values['tulumaks'], toolisemaksud, neto, maksud, tooandjamaks, tulum, tmv, aasta, pens2, tootaja2, sots2, tooandja2)
        #print(list)
        window['neto2'].update(str(list[0]))    # teksti uuendamine peale kalkuleerimist
        neto3 = list[0]
        window['pension'].update(str(list[6]))
        pension3 = list[6]
        window['tootuskindlustus2'].update(str(list[7]))
        tootuskindlustustootaja3 = list[7]
        window['tulumaks2'].update(str(list[3]))
        tulumaks3 = list[3]
        window['kokku2'].update(str(list[2]))
        kokku3 = list[2]
        window['sots2'].update(str(list[8]))
        sots3 = list[8]
        window['tooandja3'].update(str(list[9]))
        tootuskindlustustooandja3 = list[9]
        list.clear()
    #print(event, values)
    if event == 'button1':
        #direct=self.filename2
        lehenimi = str(values['kuupaevalahter1']) + "." + str(values['kuupaevalahter2'])
        #aa = "Andmed_Excelis.xlsx"
        try:
        
        #if open(aa) == True:
            
            #on vaja avada exceli fail kui on olemas ja kui ei ole, siis luua ja avada.
        
            #hea oleks saada yhele reale niimodi, et sheeti pikka teksti ei tuleks 2 korda.
            #wb = Workbook()
        
            wb = load_workbook('Andmed_Excelis.xlsx')
            #lehenimi = str(values['kuupaevalahter1']) + "." + str(values['kuupaevalahter2'])
        except:
            ws = wb.active
            #worksheet = workbook.add_worksheet(lehenimi)
            #file = open("Andmed_Excelis.xlsx", "x")
            
            #ws.title = lehenimi
            #wb.close()
            #wb = openpyxl.load_workbook('Andmed_Excelis.xlsx')
            #lehenimi = str(values['kuupaevalahter1']) + "." + str(values['kuupaevalahter2'])
        if not lehenimi in wb.sheetnames:
            #book.create_sheet(lehenimi)

            #sheet1 = wb.add_sheet('Andmed Excelis')
            #workbook = xlsxwriter.Workbook("Andmed_Excelis.xlsx")
            ws = wb.create_sheet(lehenimi)
            
        else:
            ws = wb[lehenimi]
            
        ws['A2']='Töötaja'
        ws['A3']='Tööandja'
        ws['A1']='Töötaja/Tööandja'
        ws['B1']='Eesnimi'
        ws['C1']='Perekonnanimi'
        ws['D1']='Bruto Palk'
        ws['E1']='Neto Palk'
        ws['F1']='Kogumispension'
        ws['G1']='Töötuskindlustusmaks'
        ws['H1']='Tulumaks'
        ws['I1']='Sotsiaalmaks'
        ws['J1']='Tööandja kulud kokku'
        ws['B2']=values['nimelahter1']
        ws['C2']=values['nimelahter2']
        ws['D2']=values['bruto']
        ws['E2']=neto3
        ws['F2']=pension3
        ws['G2']=tootuskindlustustootaja3
        ws['H2']=tulumaks3
        ws['I2']='-'
        ws['J2']='-'
        ws['B3']=values['nimelahter3']
        ws['C3']=values['nimelahter4']
        ws['D3']='-'
        ws['E3']='-'
        ws['F3']='-'
        ws['G3']=tootuskindlustustooandja3
        ws['H3']='-'
        ws['I3']=sots3
        ws['J3']=kokku3
        #wb.save(direct + "Andmed_Excelis.xlsx")
        wb.save(filename = 'Andmed_Excelis.xlsx')
        #book.save('Andmed_Excelis.xlsx')
            
window.close()